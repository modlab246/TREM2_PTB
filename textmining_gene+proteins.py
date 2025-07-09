#Import necessary modules from Python and standard libraries
from Bio import Entrez, Medline #for accessing and parsing PubMed records
import time
import sys
import pandas as pd    #for dataframe creation and manipulation
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

#Truncate sheet name to 31 characters
def truncate_sheet_name(name, max_length= 31):
    return name if len(name) <= max_length else name[:max_length]

#Search PubMed for tuberculosis-related articles linked to specific genes/proteins
def search_pubmed(genes_proteins):
    Entrez.email = #insert email
    Entrez.api_key = #insert API Key
    all_records = {}
    counts = {}

#Loop through each gene/protein and query PubMed    
    for item in genes_proteins:
        item_clean = item.replace('\u200b', '')
        combined_query = f'(tuberculosis[Title/Abstract]) AND ((antimicrobial[Title/Abstract]) OR (kill[Title/Abstract])) AND ({item_clean}[Title/Abstract])'
        print(f"Searching for: {combined_query}")

###Batch Processing & Rate Limiting:
Searches and downloads results for multiple genes in batch mode, with built-in delay to comply with NCBI API usage policies.

        try:
            handle = Entrez.esearch(db="pubmed", term=combined_query, retmax=5000)
            record = Entrez.read(handle)
            handle.close()
        except Exception as e:
            print(f"Error with {item}: {e}")
            continue

#Retrieve PubMed IDs and article count
        id_list = record.get("IdList", [])
        count = int(record.get("Count", "0"))
        print(f"Found {count} articles for '{item_clean}'.")

#Fetch full records if IDs are found
        if id_list:
            try:
                fetch_handle = Entrez.efetch(db="pubmed", id=id_list, rettype="medline", retmode="text")
                records = list(Medline.parse(fetch_handle))
                fetch_handle.close()
            except Exception as e:
                print(f"Fetch error for {item}: {e}")
                continue
#Filter out irrelevant articles (e.g., in vitro or serum studies) 
            filtered_records = []
            for r in records:
                abstract = r.get('AB', '').lower()
                if not any(term in abstract for term in ["in vitro", "serum"]):
                    filtered_records.append(r)

            all_records[item_clean] = filtered_records
            counts[item_clean] = len(filtered_records)
        else:
            print(f"No articles found for '{item_clean}'.")
            all_records[item_clean] = []
            counts[item_clean] = 0

        time.sleep(1) #

    return all_records, counts

#Convert retrieved MEDLINE records into structured pandas DataFrames
def create_dataframes(records):
    dataframes = {}
    for item, item_records in records.items():
        data = []
        for record in item_records:
            doi = "N/A"  #Extract DOI if available
            if 'AID' in record:
                aids = record['AID']
                doi = next((aid.split(' ')[0] for aid in aids if 'doi' in aid), "N/A")
            doi_url = f"https://doi.org/{doi}" if doi != "N/A" else "No DOI available"

#Structure each article's metadata            
            data.append({
                'PMID': record.get('PMID', 'N/A'),
                'Title': record.get('TI', 'No title available.'),
                'Authors': '; '.join(record.get('AU', ['No authors listed.'])),
                'Journal': record.get('JT', 'No journal available.'),
                'Abstract': record.get('AB', 'No abstract available.'),
                'DOI URL': doi_url
            })
        dataframes[item] = pd.DataFrame(data)
    return dataframes

#Save structured data and metadata to Excel with proper formatting
def save_to_excel(dataframes, filename, counts):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        #create a summary sheet with article counts per gene
        summary_df = pd.DataFrame([
            {"Gene/Protein": k, "Article Count": v}
            for k, v in counts.items()
        ])
        summary_df_sorted = summary_df.sort_values(by="Article Count", ascending=False)
        summary_df_sorted.to_excel(writer, sheet_name="Summary", index=False)

#Create individual sheets for each gene/protein        
        for item, df in dataframes.items():
            sheet_name = truncate_sheet_name(item.split(" OR ")[0])
            df.to_excel(writer, sheet_name=sheet_name, index=False)

#Post-process Excel to bold headers and hyperlink DOIs    
    workbook = load_workbook(filename)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
#Bold the header row
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

#Add hyperlink to DOI URL if available        
        if sheet != "Summary":
            doi_col_idx = None
            for col in worksheet.iter_cols(1, worksheet.max_column):
                if col[0].value == 'DOI URL':
                    doi_col_idx = col[0].column
                    break
            if doi_col_idx:
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=doi_col_idx, max_col=doi_col_idx):
                    cell = row[0]
                    if cell.value and cell.value != "No DOI available":
                        cell.hyperlink = cell.value
                        cell.style = 'Hyperlink'

    workbook.save(filename)
    print(f"\n Results saved to: {os.path.abspath(filename)}")

#
if __name__ == "__main__":
    genes_proteins = [
###insert gene/protein names (from DAVID output) 
    ]

 try:
        results, counts = search_pubmed(genes_proteins)
        dataframes = create_dataframes(results)
        save_to_excel(dataframes, "pubmed_antimicrobial_gene_results.xlsx", counts)
    except KeyboardInterrupt:
        print("Interrupted by user.")
        sys.exit(0)
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)
